import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import cn2an


def csv_to_xlsx(input_path, output_path):
    # 获取所有csv文件的路径
    csv_files = []
    for root, dirs, files in os.walk(input_path):
        for file in files:
            if file.endswith('.csv'):
                csv_files.append(os.path.join(root, file))

    # 逐个将csv文件转换为xlsx文件
    xlsx_files = []
    for csv_file in csv_files:
        # 构建输出文件路径
        output_file = csv_file.replace(input_path, output_path)[:-4] + '.xlsx'
        xlsx_files.append(output_file)

    return csv_files,xlsx_files


# 为数据添加index
def df_add_index(df):
    # 添加 index 列
    df.insert(0, 'index', range(1, len(df) + 1))

    ls = []

    index_i = 1

    # 遍历分组
    for index, row in df.iterrows():
        if row['省'] not in ls:
            ls.append(row['省'])
            index_i = 1
        df.loc[index, 'index'] = index_i
        index_i += 1

    return df


def csv_to_excel(file_path, outfile_name):
    # 读取csv文件数据
    df = pd.read_csv(file_path)

    # 添加index索引
    df = df_add_index(df)

    # 重排序列顺序
    df = df.reindex(columns=['index', '经销商名', "销售电话", '省', '城市'])

    # 创建Workbook对象
    wb = Workbook()
    ws = wb.active

    # 初始化行数和省、城市变量
    row = 1
    province = ''
    province_id = 1
    city = ''
    city_id = 1

    # 设置字体大小
    title_font = Font(size=18)
    font = Font(size=13)

    # 设置表格基本样式
    # 设置所有单元格的字体大小为14
    font = Font(size=14)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35

    # 遍历每一行数据
    for _, data in df.iterrows():
        # 判断省是否与前一行相同
        if province != data['省']:
            # 如果不同，则新开一行写入省名称
            province = data['省']
            province_value = '{}、{}'.format(
                cn2an.an2cn(province_id, "low"),  province)
            ws.cell(row=row, column=1, value=province_value)

            # 合并单元格
            ws.merge_cells(start_row=row, end_row=row,
                           start_column=1, end_column=3)

            # 设置居中对齐
            ws.cell(row=row, column=1).alignment = Alignment(vertical='center')

            # 设置行高
            ws.row_dimensions[row].height = 30

            # 设置字体
            ws.cell(row=row, column=1).font = title_font

            row += 1
            city = ''

            # 更新抬头数据
            province_id += 1
            city_id = 1

        # 判断城市是否与前一行相同
        if city != data['城市']:
            # 如果不同，则新开一行写入城市名称
            city = data['城市']
            city_value = '{}、{}'.format(city_id, city)

            ws.cell(row=row, column=1, value=city_value)
            ws.merge_cells(start_row=row, end_row=row,
                           start_column=1, end_column=3)
            row += 1
            city_id += 1

        # 将当前行的数据写入Excel中
        for col_idx, col_name in enumerate(df.columns):
            value = data[col_name]
            if col_idx > 2:
                break

            ws.cell(row=row, column=col_idx+1, value=value)
            ws.cell(row=row, column=col_idx +
                    1).alignment = Alignment(horizontal='center', vertical='center')

        # 合并当前行的单元格到前一行的单元格中
        for col_idx in range(3):
            if ws.cell(row=row-1, column=col_idx+1).value == ws.cell(row=row, column=col_idx+1).value:
                ws.merge_cells(start_row=row-1, end_row=row,
                               start_column=col_idx+1, end_column=col_idx+1)
        row += 1

    # 保存Excel文件
    wb.save(outfile_name)


list_csv,list_excel = csv_to_xlsx('./csv数据','./xlsx数据')


for i, j in zip(list_csv, list_excel):
    print(i,j)
    csv_to_excel(i, j)


print("数据清洗完成")
